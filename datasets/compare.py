import os
from openpyxl import Workbook
from openpyxl.styles import Alignment

# 定义要扫描的文件夹路径
folder_path_A = r'Cross_validation/fold_5/val/images'
folder_path_B = r'Cross_validation/fold_5/val/labels'

# 获取文件夹A和B中的所有文件名（不包含后缀名）
files_A = [os.path.splitext(file)[0] for file in os.listdir(folder_path_A)]
files_B = [os.path.splitext(file)[0] for file in os.listdir(folder_path_B)]

# 将文件名列表转换为字典，方便快速查找
dict_A = {file: True for file in files_A}
dict_B = {file: True for file in files_B}

# 初始化一个Workbook对象
wb = Workbook()
sheet = wb.active
sheet.title = 'File Names'

# 填入文件名到第一列和第二列
for index, file_A in enumerate(files_A, start=1):
    cell_A = sheet.cell(row=index, column=1)
    cell_A.value = file_A
    cell_A.alignment = Alignment(wrap_text=True)  # 设置第一列的文字换行

for index, file_B in enumerate(files_B, start=1):
    cell_B = sheet.cell(row=index, column=2)
    cell_B.value = file_B
    cell_B.alignment = Alignment(wrap_text=True)  # 设置第二列的文字换行

# 在第三列进行比对，并标记匹配情况
for index, file_B in enumerate(files_B, start=1):
    cell_C = sheet.cell(row=index, column=3)
    if file_B in dict_A:
        cell_C.value = "匹配"
    else:
        cell_C.value = "不匹配"

# 设置列宽
column_widths = {'A': 50, 'B': 50, 'C': 20}  # 定义列宽度字典
for column_letter, width in column_widths.items():
    sheet.column_dimensions[column_letter].width = width

# 保存工作簿
save_path = os.path.join(os.getcwd(), 'file_names_comparison.xlsx')
wb.save(save_path)

# 提示保存完成
print(f'文件名比对结果已保存到 {save_path}')
